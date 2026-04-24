"""
04_auxiliary_evidence.py
=========================
Evidenza ausiliaria a supporto dell'ipotesi H0.

Script 02 ha stabilito QUANDO i prezzi si sono rotti (lag D = tau - shock).
Script 03 ha testato SE il margine lordo e' aumentato anomalmente e con quale
robustezza (Welch t + Mann-Whitney + block perm + HAC, BH locale).

Questo script chiede tre domande complementari:

  1. VELOCITA': con quale anticipo il Brent predice i prezzi pompa?
     (Granger causality, F-test su Δlog-prezzi, lag 1-8 settimane)
     Lag significativo < 30gg e' coerente sia con mercati forward-looking
     (efficienza) sia con pricing opportunistico: da solo e' AMBIGUO.
     Classificato come evidenza esplorativa — non entra nel BH globale.

  2. ASIMMETRIA: la trasmissione e' strutturalmente asimmetrica?
     (Rockets & Feathers, GLSAR AR(1) + HAC Newey-West)
     beta_up > beta_down, p < 0.05 -> prezzi salgono piu' veloce di quanto
     scendono, indipendentemente dagli shock geopolitici.
     Anche questo e' esplorativo: l'asimmetria strutturale non implica
     speculazione contingente al singolo evento.

  3. SPECIFICITA': l'anomalia e' specifica all'Italia o comune a tutti i
     mercati EU esposti allo stesso shock wholesale?
     (Difference-in-Differences, IT vs Germania e Svezia)
     delta_DiD > 0, p < 0.05 -> eccesso italiano rispetto al benchmark.
     Il DiD controlla per la componente comune (prezzo Brent, geopolitica,
     mercato wholesale europeo) e isola la componente country-specific.
     Questo e' CONFIRMATORY su H0 e contribuisce al BH globale in script 05.

Input:
  data/dataset_merged.csv
  data/dataset_merged_with_futures.csv
  data/eu_oil_bulletin_history.xlsx

Output:
  data/granger_benzina.csv
  data/granger_diesel.csv
  data/rockets_feathers_results.csv
  data/did_results.csv
  data/auxiliary_pvalues.csv        <- input per 05_global_corrections.py
  plots/04_granger.png
  plots/04_rf.png
  plots/04_did.png
"""

import os, warnings
import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
import openpyxl as _opxl
from statsmodels.tsa.stattools import grangercausalitytests, adfuller
from statsmodels.regression.linear_model import GLSAR, OLS
from statsmodels.stats.sandwich_covariance import cov_hac
import statsmodels.api as sm
from scipy import stats

warnings.filterwarnings("ignore")
os.makedirs("data",  exist_ok=True)
os.makedirs("plots", exist_ok=True)

# ── Configurazione ─────────────────────────────────────────────────────────
ALPHA     = 0.05
MAX_LAG   = 8      # settimane per Granger
DPI       = 180

# Orizzonti per i quali testiamo H0: i due eventi con dati sufficienti.
# Hormuz escluso anche qui: finestra post-shock < 5 settimane al momento
# dell'analisi.
EVENTS = {
    "Ucraina (Feb 2022)": {
        "shock":      pd.Timestamp("2022-02-24"),
        "pre_start":  pd.Timestamp("2021-09-01"),
        "post_end":   pd.Timestamp("2022-08-31"),
    },
    "Iran-Israele (Giu 2025)": {
        "shock":      pd.Timestamp("2025-06-13"),
        "pre_start":  pd.Timestamp("2025-01-01"),
        "post_end":   pd.Timestamp("2025-10-31"),
    },
}

# Paesi di controllo per il DiD.
# Per la Svezia non ci sono distorsioni fiscali note negli anni analizzati.
# Per la Germania non ci sono sussidi al carburante negli anni analizzati.
# (La "ristourne" francese 2022 giustifica l'esclusione della Francia.)
DID_CONTROLS = {
    "Germania": ["DE", "GERMANY", "DEUTSCH"],
    "Svezia":   ["SE", "SWEDEN",  "SVERIGE"],
}

def _stars(p):
    if np.isnan(p): return ""
    return "***" if p < 0.001 else "**" if p < 0.01 else "*" if p < 0.05 else "n.s."


# ─────────────────────────────────────────────────────────────────────────────
# Carica dataset
# ─────────────────────────────────────────────────────────────────────────────
merged   = pd.read_csv("data/dataset_merged.csv",
                       index_col=0, parse_dates=True)
merged_f = pd.read_csv("data/dataset_merged_with_futures.csv",
                       index_col=0, parse_dates=True)
print(f"Dataset: {len(merged)} settimane | "
      f"{merged.index[0].date()} – {merged.index[-1].date()}\n")

# p-value confirmatory del DiD per BH globale (script 05)
aux_pvalues: list[dict] = []


# ═════════════════════════════════════════════════════════════════════════════
# §1. GRANGER CAUSALITY: Brent -> prezzi pompa
#     Evidenza esplorativa sulla velocita' di trasmissione.
# ═════════════════════════════════════════════════════════════════════════════
print("=" * 65)
print("§1. GRANGER: Brent -> prezzi pompa (velocita' di trasmissione)")
print("=" * 65)

# I livelli log sono non-stazionari (random walk): Granger su Δlog.
print("\nADF sui livelli log (ci aspettiamo non-stazionarita'):")
for col in ["log_brent", "log_benzina", "log_diesel"]:
    if col not in merged.columns:
        continue
    p_adf = adfuller(merged[col].dropna(), autolag="AIC")[1]
    staz  = "stazionario" if p_adf < ALPHA else "non stazionario -> uso Δlog"
    print(f"  {col}: p={p_adf:.4f}  [{staz}]")

merged["d_log_brent"]   = merged["log_brent"].diff()
merged["d_log_benzina"] = merged["log_benzina"].diff()
merged["d_log_diesel"]  = merged["log_diesel"].diff()

# 2020 escluso: COVID ha prodotto dinamiche anomale (WTI negativo ad aprile,
# domanda collassata ~25%) non rappresentative del meccanismo ordinario di
# trasmissione Brent -> pompa.
# Rif: Baumeister & Kilian (2020), JEP.
granger_df = merged[merged.index.year != 2020].dropna(
    subset=["d_log_brent", "d_log_benzina", "d_log_diesel"]
)
print(f"\nDati Granger: {len(granger_df)} settimane (escluso 2020 COVID)")
print()

granger_results: dict[str, pd.DataFrame] = {}

for fuel_col, fuel_name in [("d_log_benzina", "Benzina"),
                              ("d_log_diesel",  "Diesel")]:
    data2 = granger_df[[fuel_col, "d_log_brent"]].dropna()
    try:
        gc = grangercausalitytests(data2, maxlag=MAX_LAG, verbose=False)
    except Exception as e:
        print(f"  {fuel_name}: errore ({e})")
        continue

    rows = []
    for lag in range(1, MAX_LAG + 1):
        f_stat, p_val = gc[lag][0]["ssr_ftest"][:2]
        flag = " <- lag < 30gg" if p_val < ALPHA and lag * 7 < 30 else ""
        print(f"  {fuel_name} lag={lag}w ({lag*7}gg): "
              f"F={f_stat:.3f}  p={p_val:.4f} {_stars(p_val)}{flag}")
        rows.append({
            "lag_weeks":   lag,
            "lag_days":    lag * 7,
            "F_stat":      round(f_stat, 4),
            "p_value":     round(p_val, 4),
            "significant": p_val < ALPHA,
        })

    df_gc = pd.DataFrame(rows)
    granger_results[fuel_name] = df_gc
    df_gc.to_csv(f"data/granger_{fuel_name.lower()}.csv", index=False)

# Plot Granger ─────────────────────────────────────────────────────────────
if granger_results:
    def _bar_color(p):
        if p < 0.001: return "#8b1a1a"
        if p < 0.01:  return "#c0392b"
        if p < 0.05:  return "#e74c3c"
        return "#95a5a6"

    n_panels = len(granger_results)
    fig, axes = plt.subplots(1, n_panels, figsize=(7 * n_panels, 5))
    if n_panels == 1:
        axes = [axes]

    for ax, (fuel_name, df_gc) in zip(axes, granger_results.items()):
        lags_d = df_gc["lag_days"].values
        pvals  = df_gc["p_value"].values
        colors = [_bar_color(p) for p in pvals]

        ax.set_axisbelow(True)
        ax.bar(lags_d, pvals, color=colors, edgecolor="black",
               linewidth=0.6, alpha=0.90, width=5.5)
        ax.axhline(ALPHA, color="#2c3e50", lw=1.4, ls="--")
        ax.text(lags_d[-1] + 1, ALPHA + 0.005, f"alpha={ALPHA}",
                ha="right", va="bottom", fontsize=8)
        ax.axvline(30, color="#e67e22", lw=1.6, ls="--")
        ax.axvspan(0, 30, alpha=0.06, color="#e74c3c")
        ax.text(15, max(pvals) * 1.15,
                "< 30gg", ha="center", fontsize=8,
                color="#c0392b", style="italic")

        n_sig = sum(p < ALPHA and d <= 30 for p, d in zip(pvals, lags_d))
        ax.set_title(f"Granger: Brent -> {fuel_name}\n"
                     f"Lag sign. < 30gg: {n_sig}  "
                     f"(evidenza esplorativa)",
                     fontsize=10, fontweight="bold",
                     color="#b03030" if n_sig > 0 else "black")
        ax.set_xlabel("Lag (giorni)", fontsize=10)
        ax.set_ylabel("p-value (F-test)", fontsize=10)
        ax.set_xticks(lags_d)
        ax.set_xticklabels([f"{int(d)}d" for d in lags_d], fontsize=9)
        ax.set_ylim(0, max(pvals) * 1.35)
        ax.legend(handles=[
            mpatches.Patch(color="#8b1a1a", label="p<0.001 (***)"),
            mpatches.Patch(color="#c0392b", label="p<0.01  (**)"),
            mpatches.Patch(color="#e74c3c", label="p<0.05  (*)"),
            mpatches.Patch(color="#95a5a6", label="p>=0.05 (n.s.)"),
        ], fontsize=8, loc="upper right")
        ax.grid(alpha=0.3)

    fig.suptitle(
        "Granger Causality: Brent -> prezzi pompa Italia\n"
        "Misura la velocita' di trasmissione, non la speculazione diretta",
        fontsize=11, fontweight="bold", y=1.02,
    )
    fig.tight_layout(pad=1.5)
    fig.savefig("plots/04_granger.png", dpi=DPI, bbox_inches="tight")
    plt.close(fig)
    print("\n  Salvato: plots/04_granger.png")


# ═════════════════════════════════════════════════════════════════════════════
# §2. ROCKETS & FEATHERS (GLSAR AR(1) + HAC Newey-West)
#     Testa se beta_up != beta_down: asimmetria strutturale nella trasmissione.
#     Metodo: GLSAR iterativo (Cochrane-Orcutt) per rho AR(1), poi SE HAC
#     perche' i diagnostici di script 02 mostrano DW = 0.15-0.42 sistematicamente
#     -> OLS SE sono gravemente distorte verso il basso.
# ═════════════════════════════════════════════════════════════════════════════
print("\n" + "=" * 65)
print("§2. ROCKETS & FEATHERS: asimmetria strutturale beta_up vs beta_down")
print("    (evidenza esplorativa, non test diretto su H0 del margine)")
print("=" * 65)

merged["d_brent_pct"]   = merged["brent_7d_eur"].pct_change() * 100
merged["d_benzina_pct"] = merged["benzina_4w"].pct_change() * 100
merged["d_diesel_pct"]  = merged["diesel_4w"].pct_change() * 100
rf_df = merged.dropna(
    subset=["d_brent_pct", "d_benzina_pct", "d_diesel_pct"]
)

FUELS_RF = {
    "Benzina": ("d_benzina_pct", "#d6604d"),
    "Diesel":  ("d_diesel_pct",  "#31a354"),
}

rf_results: dict = {}

print(f"\nMetodo: GLSAR iterativo (Cochrane-Orcutt) + SE HAC (Newey-West, 4 lag)")
print(f"Motivazione: DW = 0.15-0.42 in tutti gli scenari -> OLS SE distorte")

for fuel_name, (dcol, fc) in FUELS_RF.items():
    y       = rf_df[dcol].values
    b_pos   = np.maximum(rf_df["d_brent_pct"].values, 0)
    b_neg   = np.minimum(rf_df["d_brent_pct"].values, 0)
    X       = np.column_stack([np.ones(len(y)), b_pos, b_neg])
    n, k    = len(y), 3

    try:
        glsar   = GLSAR(y, X, rho=1)
        gres    = glsar.iterative_fit(maxiter=10)
        b_up    = gres.params[1]
        b_down  = gres.params[2]
        rho_ar  = float(glsar.rho)
        cov_nw  = cov_hac(gres, nlags=4)
        method  = "GLSAR+HAC"
    except Exception as exc:
        print(f"  {fuel_name}: GLSAR fallito ({exc}) -> OLS+HAC")
        ols_r  = OLS(y, X).fit()
        b_up   = ols_r.params[1]
        b_down = ols_r.params[2]
        rho_ar = np.nan
        cov_nw = cov_hac(ols_r, nlags=4)
        method = "OLS+HAC"

    se_up   = float(np.sqrt(cov_nw[1, 1]))
    se_down = float(np.sqrt(cov_nw[2, 2]))
    se_diff = np.sqrt(se_up**2 + se_down**2)
    t_stat  = (b_up - b_down) / se_diff if se_diff > 0 else np.nan
    p_asym  = float(stats.t.sf(abs(t_stat), df=n - k) * 2) if not np.isnan(t_stat) else np.nan
    rf_idx  = abs(b_up) / abs(b_down) if b_down != 0 else np.inf

    rf_results[fuel_name] = {
        "b_up": b_up, "b_down": b_down,
        "se_up": se_up, "se_down": se_down,
        "rf_index": rf_idx, "t_stat": t_stat, "p_asym": p_asym,
        "rho_ar": rho_ar, "method": method, "color": fc, "dcol": dcol,
    }

    rho_s = f"{rho_ar:.3f}" if not np.isnan(rho_ar) else "N/A"
    print(f"\n  {fuel_name} [{method}]:")
    print(f"    beta_up   = {b_up:.4f}  (SE HAC={se_up:.4f})")
    print(f"    beta_down = {b_down:.4f}  (SE HAC={se_down:.4f})")
    print(f"    rho AR(1) = {rho_s}  |  R&F index = {rf_idx:.3f}")
    print(f"    p asimmetria = {p_asym:.4f} {_stars(p_asym)}")

# Salva R&F
pd.DataFrame([{
    "Carburante":   f,
    "Metodo":       r["method"],
    "b_up":         round(r["b_up"], 4),
    "SE_up_HAC":    round(r["se_up"], 4),
    "b_down":       round(r["b_down"], 4),
    "SE_down_HAC":  round(r["se_down"], 4),
    "rho_AR1":      round(r["rho_ar"], 4) if not np.isnan(r["rho_ar"]) else "N/A",
    "RF_index":     round(r["rf_index"], 3),
    "t_stat":       round(r["t_stat"], 3) if not np.isnan(r["t_stat"]) else "N/A",
    "p_asym":       round(r["p_asym"], 4) if not np.isnan(r["p_asym"]) else "N/A",
} for f, r in rf_results.items()]).to_csv("data/rockets_feathers_results.csv", index=False)

# Plot R&F ──────────────────────────────────────────────────────────────────
if rf_results:
    fig, axes = plt.subplots(1, 2, figsize=(14, 5.5), sharey=False)
    for ax, (fuel_name, res) in zip(axes, rf_results.items()):
        x_all = rf_df["d_brent_pct"].values
        y_all = rf_df[res["dcol"]].values
        valid = ~(np.isnan(x_all) | np.isnan(y_all))
        clrs  = ["#e74c3c" if b > 0 else "#3498db" for b in x_all[valid]]
        ax.scatter(x_all[valid], y_all[valid], c=clrs, alpha=0.28, s=16)

        x_up   = np.linspace(0, np.nanmax(x_all), 100)
        x_down = np.linspace(np.nanmin(x_all), 0, 100)
        ax.plot(x_up,   res["b_up"]   * x_up,   color="#e74c3c", lw=2.5,
                label=f"beta_up={res['b_up']:.4f} (Brent+)")
        ax.plot(x_down, res["b_down"] * x_down, color="#3498db", lw=2.5, ls="--",
                label=f"beta_down={res['b_down']:.4f} (Brent-)")
        ax.axhline(0, color="black", lw=0.5)
        ax.axvline(0, color="black", lw=0.5)

        p_s = f"{res['p_asym']:.4f}" if not np.isnan(res["p_asym"]) else "N/A"
        ax.set_title(f"Rockets & Feathers — {fuel_name}  [{res['method']}]\n"
                     f"R&F index={res['rf_index']:.3f}  "
                     f"p asimmetria={p_s} {_stars(res['p_asym'])}",
                     fontsize=10, fontweight="bold")
        ax.set_xlabel("Delta Brent (%)", fontsize=10)
        ax.set_ylabel(f"Delta {fuel_name} (%)", fontsize=10)
        ax.legend(fontsize=9)
        ax.grid(alpha=0.3)

    fig.suptitle(
        "Rockets & Feathers: asimmetria strutturale Brent -> pompa\n"
        "beta_up > beta_down -> prezzi salgono piu' veloce di quanto scendono",
        fontsize=11, fontweight="bold", y=1.02,
    )
    fig.tight_layout(pad=1.5)
    fig.savefig("plots/04_rf.png", dpi=DPI, bbox_inches="tight")
    plt.close(fig)
    print("\n  Salvato: plots/04_rf.png")


# ═════════════════════════════════════════════════════════════════════════════
# §3. DIFFERENCE-IN-DIFFERENCES (IT vs Germania e Svezia)
#     Testa se l'anomalia del margine e' specifica all'Italia.
#     Modello: M_{c,t} = alpha + beta1*Italy + beta2*Post + delta*(Italy*Post) + e
#     delta = stimatore DiD: variazione IT - variazione paese_controllo
#     SE HC3 (robusti a eteroschedasticita'); parallel trends test pre-shock.
#     Confirmatory: delta > 0 e significativo -> eccesso italiano specifico.
# ═════════════════════════════════════════════════════════════════════════════
print("\n" + "=" * 65)
print("§3. DIFFERENCE-IN-DIFFERENCES: anomalia specifica all'Italia?")
print("    (test confirmatory su H0 -> contribuisce al BH globale)")
print("=" * 65)

# ── Utility lettura EU Oil Bulletin con openpyxl read_only ─────────────────
# Necessario per evitare il bug data-validation nel file EU Bulletin
# (formula1 numerica nel XML causa TypeError con pd.read_excel).
EU_HIST = "data/eu_oil_bulletin_history.xlsx"

def _sheet_names(path: str) -> list[str]:
    wb = _opxl.load_workbook(path, read_only=True, data_only=True)
    ns = wb.sheetnames
    wb.close()
    return ns

def _read_sheet(path: str, sheet: str) -> pd.DataFrame:
    wb   = _opxl.load_workbook(path, read_only=True, data_only=True)
    ws   = wb[sheet]
    data = [list(r) for r in ws.iter_rows(values_only=True)]
    wb.close()
    if not data:
        return pd.DataFrame()
    hi   = next((i for i, r in enumerate(data) if any(v is not None for v in r)), 0)
    hdrs = [str(h).strip() if h is not None else f"_c{i}"
            for i, h in enumerate(data[hi])]
    df   = pd.DataFrame(data[hi + 1:], columns=hdrs)
    ic   = hdrs[0]
    df[ic] = pd.to_datetime(df[ic], errors="coerce")
    return df.set_index(ic).loc[lambda x: x.index.notna()].sort_index()

def _notax_sheet(names: list[str]) -> str:
    for s in names:
        if any(k in s.upper() for k in ["WO TAX","WITHOUT","NO TAX","NOTAX"]):
            return s
    return names[1] if len(names) > 1 else names[0]

# ── Carica margini per i paesi di controllo ────────────────────────────────
CONTROL_MARGINS: dict[str, pd.DataFrame] = {}

# Wholesale costs (stessi per IT e paesi controllo: mercato spot europeo)
eurobob_al = merged_f["eurobob_eur_l"] if "eurobob_eur_l" in merged_f.columns else None
gasoil_al  = merged_f["gasoil_eur_l"]  if "gasoil_eur_l"  in merged_f.columns else None

try:
    snames  = _sheet_names(EU_HIST)
    notax   = _notax_sheet(snames)
    df_eu   = _read_sheet(EU_HIST, notax).apply(pd.to_numeric, errors="coerce")

    for paese, prefissi in DID_CONTROLS.items():
        all_c = [c for c in df_eu.columns if any(
            str(c).upper().startswith(p) or p in str(c).upper()
            for p in prefissi
        )]
        b_c = [c for c in all_c if any(k in str(c).lower()
               for k in ["95","benz","petrol","gasol","euro","unleaded"])]
        d_c = [c for c in all_c if any(k in str(c).lower()
               for k in ["diesel","gas_oil","gasoil"])]
        if not b_c and all_c: b_c = [all_c[0]]
        if not d_c and len(all_c) >= 2: d_c = [all_c[1]]

        if not b_c or not d_c:
            print(f"  {paese}: colonne non trovate nel foglio '{notax}' — skip")
            continue

        pump = pd.concat([df_eu[b_c[0]].rename("benzina_eur_l"),
                          df_eu[d_c[0]].rename("diesel_eur_l")], axis=1)
        pump = pump[pump.index >= "2019-01-01"].dropna(how="all")
        pump = pump.resample("W-MON").mean()

        # Normalizzazione unita' (EU Bulletin a volte pubblica in EUR/1000L)
        for col in ["benzina_eur_l", "diesel_eur_l"]:
            if col in pump.columns:
                med = pump[col].dropna().median()
                pump[col] = pump[col] / (1000.0 if med > 10 else 1.0)

        # Margini = prezzi pompa paese - costo wholesale europeo (stessa proxy IT)
        if eurobob_al is not None:
            pump["margine_benz_crack"] = (
                pump["benzina_eur_l"] - eurobob_al.reindex(pump.index).ffill(limit=4)
            )
        if gasoil_al is not None:
            pump["margine_dies_crack"] = (
                pump["diesel_eur_l"] - gasoil_al.reindex(pump.index).ffill(limit=4)
            )

        CONTROL_MARGINS[paese] = pump
        n_ok = pump.dropna(subset=["benzina_eur_l"]).shape[0]
        print(f"  {paese}: {n_ok} settimane  |  "
              f"colonne: {b_c[0]}, {d_c[0]}")

except FileNotFoundError:
    print(f"  {EU_HIST} non trovato — DiD saltato.")
    print(f"  Eseguire prima 01_data_pipeline.py.")
except Exception as exc:
    print(f"  Errore caricamento dati controllo: {exc}")

# ── Margini IT (gia' calcolati in script 01/03) ────────────────────────────
IT_MARGIN = {
    "Benzina": "margine_benz_crack",
    "Diesel":  "margine_dies_crack",
}
# Verifica disponibilita'
it_margins_ok = all(col in merged_f.columns for col in IT_MARGIN.values())
if not it_margins_ok:
    print("  Margini IT non trovati in dataset_merged_with_futures.csv")
    print("  Eseguire prima 01_data_pipeline.py.")

# ── Stima DiD per ogni evento x paese x carburante ────────────────────────
did_rows = []

if CONTROL_MARGINS and it_margins_ok:
    for paese, ctrl in CONTROL_MARGINS.items():
        for ev_name, cfg in EVENTS.items():
            shock     = cfg["shock"]
            pre_start = cfg["pre_start"]
            post_end  = cfg["post_end"]

            for fuel_name, it_col in IT_MARGIN.items():
                ct_col = it_col  # stesso nome, DataFrame diverso

                if ct_col not in ctrl.columns:
                    continue

                it_pre  = merged_f.loc[pre_start:shock,   it_col].dropna()
                it_post = merged_f.loc[shock:post_end,     it_col].dropna()
                ct_pre  = ctrl.loc[pre_start:shock,  ct_col].dropna()
                ct_post = ctrl.loc[shock:post_end,   ct_col].dropna()

                if any(len(s) < 3 for s in [it_pre, it_post, ct_pre, ct_post]):
                    continue

                # Parallel Trends Test (pre-shock):
                # H0: l'andamento pre-shock IT e CT e' parallelo.
                # Testa il coefficiente di interazione Italy*t in una regressione
                # lineare sul solo periodo pre-shock.
                pt_pvalue = np.nan
                pt_valida = None
                try:
                    all_pre_idx = it_pre.index.union(ct_pre.index).sort_values()
                    t_it = np.array([(d - all_pre_idx[0]).days
                                     for d in it_pre.index], dtype=float)
                    t_ct = np.array([(d - all_pre_idx[0]).days
                                     for d in ct_pre.index], dtype=float)
                    rows_pt = (
                        [(1, t, v) for t, v in zip(t_it, it_pre.values)] +
                        [(0, t, v) for t, v in zip(t_ct, ct_pre.values)]
                    )
                    df_pt = pd.DataFrame(rows_pt, columns=["Italy", "t", "M"])
                    df_pt["Ixt"] = df_pt["Italy"] * df_pt["t"]
                    X_pt   = sm.add_constant(df_pt[["Italy","t","Ixt"]].values)
                    ols_pt = sm.OLS(df_pt["M"].values, X_pt).fit(cov_type="HC3")
                    pt_pvalue = float(ols_pt.pvalues[3])
                    pt_valida = bool(pt_pvalue >= ALPHA)
                except Exception:
                    pass  # PTA non calcolabile, procediamo comunque

                # Modello DiD:
                # M_{c,t} = alpha + beta1*Italy + beta2*Post + delta*(Italy*Post) + e
                rows_panel = (
                    [(1, 0, v) for v in it_pre.values]  +
                    [(1, 1, v) for v in it_post.values] +
                    [(0, 0, v) for v in ct_pre.values]  +
                    [(0, 1, v) for v in ct_post.values]
                )
                df_panel = pd.DataFrame(rows_panel, columns=["Italy","Post","M"])
                df_panel["IxP"] = df_panel["Italy"] * df_panel["Post"]
                X_did = sm.add_constant(df_panel[["Italy","Post","IxP"]].values)

                try:
                    ols_did  = sm.OLS(df_panel["M"].values, X_did).fit(cov_type="HC3")
                    delta    = float(ols_did.params[3])
                    se_delta = float(ols_did.bse[3])
                    p_did    = float(ols_did.pvalues[3])
                    ci_lo    = delta - 1.96 * se_delta
                    ci_hi    = delta + 1.96 * se_delta
                except Exception as exc:
                    print(f"  DiD errore {paese}|{ev_name}|{fuel_name}: {exc}")
                    continue

                pta_flag = (
                    " [PTA valida]" if pt_valida is True
                    else " [PTA violata -- interpretare con cautela]" if pt_valida is False
                    else " [PTA non calcolata]"
                )
                print(f"\n  [{paese}] {ev_name.split('(')[0].strip()} | {fuel_name}:")
                print(f"    delta_DiD = {delta:+.4f} EUR/l  "
                      f"SE={se_delta:.4f}  "
                      f"p={p_did:.4f} {_stars(p_did)}")
                print(f"    CI 95%: [{ci_lo:+.4f}, {ci_hi:+.4f}]")
                if not np.isnan(pt_pvalue):
                    print(f"    PTA: p={pt_pvalue:.3f}{pta_flag}")

                did_rows.append({
                    "Evento":          ev_name,
                    "Paese_controllo": paese,
                    "Carburante":      fuel_name,
                    "n_IT_pre":        len(it_pre),
                    "n_IT_post":       len(it_post),
                    "n_CT_pre":        len(ct_pre),
                    "n_CT_post":       len(ct_post),
                    "PTA_pvalue":      round(pt_pvalue, 4) if not np.isnan(pt_pvalue) else "N/A",
                    "PTA_non_rigettata": pt_valida,
                    "delta_DiD":       round(delta, 4),
                    "SE_HC3":          round(se_delta, 4),
                    "CI_95_lo":        round(ci_lo, 4),
                    "CI_95_hi":        round(ci_hi, 4),
                    "p_value":         round(p_did, 6),
                    "H0":              "RIFIUTATA" if p_did < ALPHA else "non rifiutata",
                })

                # DiD e' confirmatory su H0 -> va nel BH globale
                aux_pvalues.append({
                    "fonte":       f"DiD_{ev_name}_{paese}_{fuel_name}",
                    "tipo":        "confirmatory",
                    "descrizione": f"{ev_name} | IT vs {paese} | {fuel_name}",
                    "p_value":     p_did,
                })

# Salva DiD
if did_rows:
    pd.DataFrame(did_rows).to_csv("data/did_results.csv", index=False)
    print(f"\n  Salvato: data/did_results.csv ({len(did_rows)} stime)")

    # Plot DiD: delta con CI 95% per ogni combinazione
    df_did  = pd.DataFrame(did_rows)
    fig_d, ax_d = plt.subplots(figsize=(12, max(5, len(df_did) * 0.70)))
    labels_d = [
        f"{r['Evento'].split('(')[0].strip()} | {r['Paese_controllo']} | {r['Carburante']}"
        for _, r in df_did.iterrows()
    ]
    delta_d = df_did["delta_DiD"].values
    ci_lo_d = df_did["delta_DiD"].values - df_did["CI_95_lo"].values
    ci_hi_d = df_did["CI_95_hi"].values - df_did["delta_DiD"].values
    clr_d   = ["#e74c3c" if p < ALPHA else "#95a5a6"
                for p in df_did["p_value"].values]

    ax_d.barh(range(len(df_did)), delta_d, color=clr_d,
              edgecolor="black", lw=0.5, alpha=0.85)
    for i, (lo, hi) in enumerate(zip(ci_lo_d, ci_hi_d)):
        ax_d.errorbar(delta_d[i], i, xerr=[[lo],[hi]],
                      fmt="none", color="black", capsize=5, lw=1.5)
    ax_d.axvline(0, color="black", lw=1.0, ls="--")
    ax_d.set_yticks(range(len(df_did)))
    ax_d.set_yticklabels(labels_d, fontsize=8.5)
    ax_d.set_xlabel("delta DiD (EUR/litro, CI 95% HC3)", fontsize=10)
    ax_d.set_title(
        "Difference-in-Differences: eccesso margine Italia vs paesi controllo\n"
        "rosso = p < 0.05  |  CI 95% HC3  |  "
        "delta > 0 -> anomalia specifica IT",
        fontsize=10, fontweight="bold",
    )
    ax_d.grid(alpha=0.3, axis="x")
    plt.tight_layout(pad=1.5)
    fig_d.savefig("plots/04_did.png", dpi=DPI, bbox_inches="tight")
    plt.close(fig_d)
    print("  Salvato: plots/04_did.png")
else:
    print("  DiD non eseguito (dati controllo non disponibili).")

# ─────────────────────────────────────────────────────────────────────────────
# Salva p-value ausiliari per BH globale (solo DiD, il resto e' esplorativo)
# ─────────────────────────────────────────────────────────────────────────────
pd.DataFrame(aux_pvalues).to_csv("data/auxiliary_pvalues.csv", index=False)
print(f"\nSalvato: data/auxiliary_pvalues.csv ({len(aux_pvalues)} test confirmatory DiD)")

print("\nScript 04 completato.")
print("  Granger:  esplorativo (velocita' trasmissione) -> NON nel BH globale")
print("  R&F:      esplorativo (asimmetria strutturale)  -> NON nel BH globale")
print("  DiD:      confirmatory (specificita' IT)        -> data/auxiliary_pvalues.csv")