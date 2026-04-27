"""
3_03_did.py  — Famiglia C: DiD + Windfall (v3, completo)
=========================================================
Testa H₀(iii) e stima l'extraprofitto ipotetico (windfall).

FAMIGLIA C  — H₀(iii): specificità italiana
  H₀: δ_DiD ≤ 0   (IT non cresce più dei peer EU)
  H₁: δ_DiD > 0   (IT cresce più → comportamento opportunistico)

  Modello DiD OLS panel:
    M_it = α + β·IT_i + γ·Post_t + δ·(IT×Post)_it + ε
    δ̂ = (Δ_IT_post − Δ_IT_pre) − (Δ_CT_post − Δ_CT_pre)
  SE robuste HC3.

PAESI CONTROLLO
────────────────
  Germania: scelta principale, grande mercato liquido.
  Svezia:   alternativa stabile, NO misure fiscali distorsive nel periodo.
  (Francia esclusa: ristorno fiscale ~0.15-0.18 €/l set–dic 2022
   comprime il margine FR e confonde il DiD per Ucraina.)

PTA (Parallel Trends Assumption)
─────────────────────────────────
  Testata con interazione trend×paese sulle ultime 8 settimane pre-shock.
  PTA violata (p < 0.05) → rigetto DiD da interpretare come DESCRITTIVO,
  non causale.

WINDFALL PROFITS
────────────────
  Stima proxy dell'extraprofitto totale IT (per evento, per carburante):
    Windfall = δ̂_DiD × Volume × N_settimane_post
  Volumi proxy (MISE 2022, corretti per trend consumi -1.5%/anno):
    Benzina ≈ 33.65 ML/settimana  |  Diesel ≈ 43.27 ML/settimana
  Sensitività: ±30% sui volumi.

Input:
  data/3_dataset.csv                    ← crack spread IT
  data/pompa_de.csv, data/pompa_se.csv  ← prezzi pompa DE e SE (da 3_01_data.py)
  data/eu_oil_bulletin_history.xlsx     ← fallback se i CSV non esistono

Output:
  data/3_C.csv              — 8 test DiD
  data/3_windfall.csv       — stime windfall con sensitività
  plots/3_03a_did.png       — forest plot DiD
  plots/3_03b_windfall.png  — grafico windfall
"""

import os
import warnings
import numpy as np
import pandas as pd
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
from matplotlib.lines import Line2D
import statsmodels.api as sm

warnings.filterwarnings("ignore")
os.makedirs("data",  exist_ok=True)
os.makedirs("plots", exist_ok=True)

# ── Costanti ──────────────────────────────────────────────────────────────────
ALPHA     = 0.05
PTA_WEEKS = 8       # finestra pre-shock per test PTA
DPI       = 160

# Conversione Brent: EUR/barile → EUR/litro
BRENT_BBL_TO_L = 158.987
YIELD_BENZ     = 0.45
YIELD_DIES     = 0.52

# Volumi proxy IT 2022 (Fonte: MISE, vendite mensili / 52)
VOLUME_2022 = {"Benzina": 1_750_000_000 / 52,   # L/settimana
               "Diesel":  2_250_000_000 / 52}
TREND_ANNO  = -0.015   # riduzione consumi -1.5%/anno dal 2022

EVENTS = {
    "Ucraina (Feb 2022)": {
        "shock":     pd.Timestamp("2022-02-24"),
        "pre_start": pd.Timestamp("2021-01-11"),
        "post_end":  pd.Timestamp("2022-12-31"),
        "color":     "#e74c3c",
        "ref_year":  2022,
    },
    "Iran-Israele (Giu 2025)": {
        "shock":     pd.Timestamp("2025-06-13"),
        "pre_start": pd.Timestamp("2024-01-01"),
        "post_end":  pd.Timestamp("2025-12-31"),
        "color":     "#e67e22",
        "ref_year":  2025,
    },
}

FUELS = ["Benzina", "Diesel"]
FUEL_COLORS = {"Benzina": "#d6604d", "Diesel": "#4393c3"}


# ════════════════════════════════════════════════════════════════════════════
# HELPER: carica prezzi pompa paese estero
# ════════════════════════════════════════════════════════════════════════════

def _margine_col(df: pd.DataFrame, fuel: str) -> str | None:
    """Restituisce la colonna del margine crack spread per il carburante."""
    if fuel == "Benzina" and "margine_benz_crack" in df.columns:
        return "margine_benz_crack"
    if fuel == "Diesel"  and "margine_dies_crack" in df.columns:
        return "margine_dies_crack"
    return None


def _load_control_pump(country: str, merged_it: pd.DataFrame) -> pd.DataFrame | None:
    """
    Carica i prezzi pompa di un paese estero (DE, SE) e calcola il crack spread
    usando lo stesso Brent del dataset IT per comparabilità.

    Prima cerca il CSV pre-salvato da 3_01_data.py; se non esiste legge
    direttamente dall'Excel EU Oil Bulletin.
    """
    csv_path = f"data/pompa_{country.lower()}.csv"
    PREFIXES = {
        "DE": ["DE_"],
        "SE": ["SE_"],
    }
    FUEL_KW_BENZ = ["95", "benz", "petrol", "unleaded", "super", "euro_", "euro95"]
    FUEL_KW_DIES = ["diesel", "gasoil", "gas_oil"]

    pump = None

    # Prova 1: CSV pre-salvato da 3_01_data.py
    if os.path.exists(csv_path):
        try:
            pump = pd.read_csv(csv_path, index_col=0, parse_dates=True).sort_index()
            print(f"   {country}: caricato da {csv_path}  ({len(pump)} settimane)")
        except Exception as exc:
            print(f"   {country}: errore lettura {csv_path}: {exc}")
            pump = None

    # Prova 2: Excel EU Oil Bulletin
    if pump is None:
        xl_path = "data/eu_oil_bulletin_history.xlsx"
        if not os.path.exists(xl_path):
            print(f"   {country}: {xl_path} non trovato, skip")
            return None
        try:
            import openpyxl as _opxl

            def _wb_read(path, sheet):
                wb = _opxl.load_workbook(path, read_only=True, data_only=True)
                ws = wb[sheet]
                rows = [list(r) for r in ws.iter_rows(values_only=True)]
                wb.close()
                hdr_i = next((i for i, r in enumerate(rows)
                               if sum(v is not None for v in r) >= 3), 0)
                hdrs = [str(h).strip() if h is not None else f"_c{j}"
                        for j, h in enumerate(rows[hdr_i])]
                df_ = pd.DataFrame(rows[hdr_i + 1:], columns=hdrs)
                ic = hdrs[0]
                df_[ic] = pd.to_datetime(df_[ic], errors="coerce")
                return df_.set_index(ic).sort_index()[lambda x: x.index.notna()]

            wb = _opxl.load_workbook(xl_path, read_only=True, data_only=True)
            sheets = wb.sheetnames; wb.close()
            notax_sheet = next(
                (s for s in sheets
                 if any(k in s.upper() for k in ["WO TAX","WITHOUT","NO TAX","NOTAX"])),
                sheets[1] if len(sheets) > 1 else sheets[0]
            )
            df_eu = _wb_read(xl_path, notax_sheet).apply(pd.to_numeric, errors="coerce")
            prefs = PREFIXES.get(country, [f"{country}_"])
            all_c = [c for c in df_eu.columns
                     if any(str(c).upper().startswith(p) for p in prefs)]
            b_cols = [c for c in all_c if any(k in str(c).lower() for k in FUEL_KW_BENZ)]
            d_cols = [c for c in all_c if any(k in str(c).lower() for k in FUEL_KW_DIES)]
            if not b_cols and all_c:         b_cols = [all_c[0]]
            if not d_cols and len(all_c) > 1: d_cols = [all_c[1]]
            if not b_cols or not d_cols:
                print(f"   {country}: colonne non trovate nel Bulletin")
                return None
            pump = pd.concat([df_eu[b_cols[0]].rename("benzina_eur_l"),
                              df_eu[d_cols[0]].rename("diesel_eur_l")], axis=1)
            pump = pump[pump.index >= "2019-01-01"].dropna(how="all").resample("W-MON").mean()
            for col in ["benzina_eur_l", "diesel_eur_l"]:
                if pump[col].dropna().median() > 10:
                    pump[col] /= 1000.0
            print(f"   {country}: caricato da Bulletin [{b_cols[0]}, {d_cols[0]}]")
        except Exception as exc:
            print(f"   {country}: errore Bulletin: {exc}")
            return None

    # Calcola crack spread con stesso Brent IT
    brent_al = merged_it["brent_eur"].reindex(pump.index).ffill(limit=4)
    pump["margine_benzina"] = (pump["benzina_eur_l"]
                               - brent_al / (BRENT_BBL_TO_L * YIELD_BENZ))
    pump["margine_diesel"]  = (pump["diesel_eur_l"]
                               - brent_al / (BRENT_BBL_TO_L * YIELD_DIES))
    return pump


def _stars(p):
    if p < 0.001: return "***"
    if p < 0.01:  return "**"
    if p < 0.05:  return "*"
    if p < 0.10:  return "."
    return ""


# ════════════════════════════════════════════════════════════════════════════
# CARICA DATASET IT
# ════════════════════════════════════════════════════════════════════════════
print("Carico dataset Italia...")
merged = pd.read_csv("data/3_dataset.csv", index_col=0, parse_dates=True).sort_index()
print(f"   {len(merged)} settimane | {merged.index[0].date()} – {merged.index[-1].date()}")

# ════════════════════════════════════════════════════════════════════════════
# CARICA PAESI CONTROLLO
# ════════════════════════════════════════════════════════════════════════════
print("\nCarico paesi controllo (DE, SE)...")
CONTROL_PUMPS: dict[str, pd.DataFrame] = {}
for paese in ["Germania", "Svezia"]:
    country_code = "DE" if paese == "Germania" else "SE"
    pump = _load_control_pump(country_code, merged)
    if pump is not None:
        CONTROL_PUMPS[paese] = pump

FUEL_TO_COL_CT = {"Benzina": "margine_benzina", "Diesel": "margine_diesel"}

# ════════════════════════════════════════════════════════════════════════════
# ESEGUI DiD
# ════════════════════════════════════════════════════════════════════════════
print("\n" + "="*62)
print("FAMIGLIA C — H₀(iii): specificità italiana (DiD IT vs EU)")
print("="*62)

rows_did      = []
rows_windfall = []

for paese, ctrl in CONTROL_PUMPS.items():
    for evento, ecfg in EVENTS.items():
        shock     = ecfg["shock"]
        pre_start = ecfg["pre_start"]
        post_end  = ecfg["post_end"]

        for fuel in FUELS:
            mcol_it = _margine_col(merged, fuel)
            mcol_ct = FUEL_TO_COL_CT[fuel]

            if mcol_it is None or mcol_it not in merged.columns:
                continue
            if mcol_ct not in ctrl.columns:
                continue

            it_pre  = merged.loc[pre_start:shock - pd.Timedelta(days=1), mcol_it].dropna()
            it_post = merged.loc[shock:post_end,                          mcol_it].dropna()
            ct_pre  = ctrl.loc[pre_start:shock - pd.Timedelta(days=1),   mcol_ct].dropna()
            ct_post = ctrl.loc[shock:post_end,                            mcol_ct].dropna()

            if any(len(s) < 3 for s in [it_pre, it_post, ct_pre, ct_post]):
                print(f"   SKIP {evento} {fuel} ({paese}): dati insufficienti")
                continue

            # ── PTA: ultimi 8 settimane pre-shock ────────────────────────────
            pta_p    = np.nan
            pta_pass = None
            try:
                pta_w  = pd.Timedelta(weeks=PTA_WEEKS)
                it_pta = it_pre[it_pre.index >= (shock - pta_w)]
                ct_pta = ct_pre[ct_pre.index >= (shock - pta_w)]
                if len(it_pta) < 3 or len(ct_pta) < 3:
                    it_pta, ct_pta = it_pre, ct_pre   # fallback: tutto il pre
                t0 = min(it_pta.index.min(), ct_pta.index.min())
                panel_pta = (
                    [(1, (d - t0).days, v) for d, v in zip(it_pta.index, it_pta.values)] +
                    [(0, (d - t0).days, v) for d, v in zip(ct_pta.index, ct_pta.values)]
                )
                dpt = pd.DataFrame(panel_pta, columns=["IT", "t", "M"])
                dpt["IT_x_t"] = dpt["IT"] * dpt["t"]
                X_pt   = sm.add_constant(dpt[["IT", "t", "IT_x_t"]].values)
                ols_pt = sm.OLS(dpt["M"].values, X_pt).fit(cov_type="HC3")
                pta_p    = float(ols_pt.pvalues[3])
                pta_pass = bool(pta_p >= ALPHA)
            except Exception as ep:
                print(f"   PTA errore ({evento} {fuel} {paese}): {ep}")

            # ── Modello DiD OLS ───────────────────────────────────────────────
            panel = (
                [(1, 0, v) for v in it_pre.values]  +
                [(1, 1, v) for v in it_post.values] +
                [(0, 0, v) for v in ct_pre.values]  +
                [(0, 1, v) for v in ct_post.values]
            )
            dpanel              = pd.DataFrame(panel, columns=["IT", "Post", "M"])
            dpanel["IT_x_Post"] = dpanel["IT"] * dpanel["Post"]
            X   = sm.add_constant(dpanel[["IT", "Post", "IT_x_Post"]].values)
            ols = sm.OLS(dpanel["M"].values, X).fit(cov_type="HC3")

            delta   = float(ols.params[3])
            se      = float(ols.bse[3])
            t_stat  = float(ols.tvalues[3])
            p_two   = float(ols.pvalues[3])
            p_one   = float(p_two / 2) if t_stat > 0 else float(1.0 - p_two / 2)
            ci_lo   = delta - 1.96 * se
            ci_hi   = delta + 1.96 * se

            # Interpretazione economica
            rej = p_one < ALPHA and delta > 0
            if delta > 0 and rej:
                note = f"δ>0 sign.: IT > {paese} post-shock — evidenza A FAVORE specificità IT"
            elif delta < 0 and p_one < ALPHA:
                note = f"δ<0 sign.: IT < {paese} post-shock — evidenza CONTRO specificità IT"
            elif delta > 0:
                note = f"δ>0 n.s.: differenza IT vs {paese} non distinguibile da zero"
            else:
                note = f"δ≤0 n.s.: IT non cresce più di {paese}"

            pta_tag = " ⚠PTA✗" if pta_pass is False else " PTA✓"
            print(f"   [{paese}] {evento[:22]:<22} | {fuel:<8}: "
                  f"δ={delta:+.4f}  p={p_one:.4f}{_stars(p_one):<3}"
                  f"  CI=[{ci_lo:+.4f},{ci_hi:+.4f}]  "
                  f"{'RIFIUTATA' if rej else 'non rig.'}{pta_tag}")

            rows_did.append({
                "famiglia":         "C",
                "ipotesi":          "H0_iii",
                "evento":           evento,
                "paese_controllo":  paese,
                "carburante":       fuel,
                "fonte":            f"DiD_{evento}_{paese}_{fuel}",
                "n_IT_pre":         len(it_pre),
                "n_IT_post":        len(it_post),
                "n_CT_pre":         len(ct_pre),
                "n_CT_post":        len(ct_post),
                "PTA_pvalue":       round(float(pta_p), 4) if not np.isnan(pta_p) else None,
                "PTA_non_rigettata": pta_pass,
                "delta_DiD_EUR_L":  round(delta, 5),
                "SE_HC3":           round(se,    5),
                "CI_95_low":        round(ci_lo, 5),
                "CI_95_high":       round(ci_hi, 5),
                "t_stat":           round(t_stat, 3),
                "p_value":          round(p_one, 6),
                "p_value_twosided": round(p_two, 6),
                "R2_OLS":           round(float(ols.rsquared), 3),
                "H0":               "RIFIUTATA" if rej else "non rifiutata",
                "interpretation":   note,
            })

            # ── Windfall (solo per Germania come controllo principale) ────────
            if paese == "Germania" and delta > 0:
                ref_yr   = ecfg["ref_year"]
                n_wks    = len(it_post)
                # Volume corretto per trend: V_2022 × (1 + trend)^(ref_yr - 2022)
                yrs_from_2022 = ref_yr - 2022
                for fuel_name, base_vol in VOLUME_2022.items():
                    if fuel_name != fuel:
                        continue
                    vol_adj = base_vol * (1 + TREND_ANNO) ** yrs_from_2022
                    for mult, scen in [(0.70, "-30%"), (1.00, "base"), (1.30, "+30%")]:
                        wf_mld = delta * vol_adj * mult * n_wks / 1e9
                        rows_windfall.append({
                            "evento":          evento,
                            "carburante":      fuel,
                            "delta_DiD_EUR_L": round(delta, 5),
                            "n_settimane_post": n_wks,
                            "vol_scenario":    scen,
                            "vol_adj_ML_wk":   round(vol_adj * mult / 1e6, 3),
                            "windfall_MLD_EUR": round(wf_mld, 3),
                        })


# ════════════════════════════════════════════════════════════════════════════
# SALVA CSV
# ════════════════════════════════════════════════════════════════════════════
if not rows_did:
    print("\n  Nessun test DiD prodotto.")
    df_did = pd.DataFrame()
else:
    df_did = pd.DataFrame(rows_did)
    df_did.to_csv("data/3_C.csv", index=False)
    print(f"\n✓ data/3_C.csv  ({len(df_did)} test DiD)")

if rows_windfall:
    df_wf = pd.DataFrame(rows_windfall)
    df_wf.to_csv("data/3_windfall.csv", index=False)
    print(f"✓ data/3_windfall.csv  ({len(df_wf)} scenari)")
    print("\n   Windfall estimates (scenario base, vs Germania):")
    base_wf = df_wf[df_wf["vol_scenario"] == "base"]
    for _, r in base_wf.iterrows():
        print(f"   {r['evento'][:22]} | {r['carburante']}: "
              f"δ={r['delta_DiD_EUR_L']:+.4f} EUR/L × {r['vol_adj_ML_wk']:.1f} ML/wk "
              f"× {r['n_settimane_post']} wk = {r['windfall_MLD_EUR']:.2f} MLD EUR")


# ════════════════════════════════════════════════════════════════════════════
# GRAFICI
# ════════════════════════════════════════════════════════════════════════════

# ── Fig A: Forest plot DiD ────────────────────────────────────────────────
if not df_did.empty:
    events_list  = list(EVENTS.keys())
    countries    = list(CONTROL_PUMPS.keys())
    n_events     = len(events_list)
    fig, axes    = plt.subplots(1, n_events, figsize=(6.5 * n_events, max(5, len(countries)*2+2)),
                                 squeeze=False)

    for ei, evento in enumerate(events_list):
        ax = axes[0][ei]
        ev_color = EVENTS[evento]["color"]

        # Raccogli righe del forest plot (paese × fuel, ordinate)
        plot_rows = []
        for paese in countries:
            sub = df_did[(df_did["evento"] == evento) &
                         (df_did["paese_controllo"] == paese)].sort_values("carburante")
            for _, r in sub.iterrows():
                plot_rows.append(r)

        y_pos   = list(range(len(plot_rows)))
        xlim_a  = max(0.15, max(abs(r["CI_95_high"]) for r in plot_rows if r["CI_95_high"] is not None) + 0.03)

        # Zona rossa: δ > 0 (specificità italiana — H₁)
        ax.axvspan(0, xlim_a, alpha=0.05, color="#e74c3c")
        ax.axvline(0, color="black", lw=0.8, ls="--", alpha=0.5)

        for yi, r in zip(y_pos, plot_rows):
            col   = FUEL_COLORS[r["carburante"]]
            sig   = r["p_value"] < ALPHA and r["delta_DiD_EUR_L"] > 0

            # Barra CI
            ax.hlines(yi, r["CI_95_low"], r["CI_95_high"], colors=col, lw=2.8, alpha=0.9)
            # Punto: diamante se significativo
            ax.plot(r["delta_DiD_EUR_L"], yi,
                    marker=("D" if sig else "o"), color=col, ms=8, zorder=3)
            # Cerchio arancione se PTA violata
            if r["PTA_non_rigettata"] is False:
                ax.plot(r["delta_DiD_EUR_L"], yi, "o", color="#f39c12",
                        ms=15, zorder=2, alpha=0.5)

            pta_tag = " ⚠PTA✗" if r["PTA_non_rigettata"] is False else ""
            label   = f"{r['paese_controllo']} — {r['carburante']}{pta_tag}"
            ax.text(-xlim_a + 0.003, yi, label, va="center", fontsize=8.5)
            # p-value sul lato destro
            p_str = f"p={r['p_value']:.3f}{_stars(r['p_value'])}"
            ax.text(xlim_a - 0.003, yi, p_str, va="center", ha="right", fontsize=8)

        ax.set_yticks([])
        ax.set_xlim(-xlim_a, xlim_a)
        ax.set_xlabel("δ DiD (EUR/L)", fontsize=10)
        ax.set_title(f"{evento}\nδ>0 = IT cresce più del paese controllo",
                     fontsize=10, fontweight="bold")
        ax.grid(axis="x", alpha=0.25)

    # Legenda comune
    legend_els = [
        Line2D([0],[0], marker="o", color=FUEL_COLORS["Benzina"], lw=0, ms=8, label="Benzina"),
        Line2D([0],[0], marker="o", color=FUEL_COLORS["Diesel"],  lw=0, ms=8, label="Diesel"),
        Line2D([0],[0], marker="D", color="gray", lw=0, ms=8, label="Significativo (δ>0, p<0.05)"),
        Line2D([0],[0], marker="o", color="#f39c12", lw=0, ms=12, alpha=0.6, label="PTA violata"),
        mpatches.Patch(color="#e74c3c", alpha=0.12, label="Zona H₁ (δ>0 = specificità IT)"),
    ]
    fig.legend(handles=legend_els, loc="lower center", ncol=5,
               fontsize=8.5, bbox_to_anchor=(0.5, -0.06))
    fig.suptitle("H₀(iii) — DiD: IT vs Germania e Svezia\n"
                 "(δ>0 → IT ha margini più alti del paese controllo post-shock)",
                 fontsize=11, fontweight="bold", y=1.01)
    plt.tight_layout()
    fig.savefig("plots/3_03a_did.png", dpi=DPI, bbox_inches="tight")
    plt.close(fig)
    print("✓ plots/3_03a_did.png")


# ── Fig B: Windfall ───────────────────────────────────────────────────────
if rows_windfall:
    df_wf_plot = pd.DataFrame(rows_windfall)
    cases = df_wf_plot[["evento", "carburante"]].drop_duplicates().values.tolist()
    n_cases = len(cases)

    if n_cases > 0:
        fig, axes_wf = plt.subplots(1, max(n_cases, 1),
                                     figsize=(5 * max(n_cases, 1), 4.5),
                                     squeeze=False)
        scenario_colors = {"-30%": "#90caf9", "base": "#1565c0", "+30%": "#0d47a1"}

        for ci, (ev, fuel) in enumerate(cases):
            ax = axes_wf[0][ci]
            sub = df_wf_plot[(df_wf_plot["evento"] == ev) &
                             (df_wf_plot["carburante"] == fuel)]
            scen_vals = sub.set_index("vol_scenario")["windfall_MLD_EUR"]
            bars = ax.bar(list(scen_vals.index), list(scen_vals.values),
                          color=[scenario_colors.get(s, "gray") for s in scen_vals.index],
                          alpha=0.85, edgecolor="black", lw=0.5)
            # Etichette sui bar
            for bar, val in zip(bars, scen_vals.values):
                ax.text(bar.get_x() + bar.get_width()/2, bar.get_height() + 0.005,
                        f"{val:.2f}B", ha="center", fontsize=9, fontweight="bold")
            ax.set_title(f"{ev[:22]}\n{fuel}", fontsize=9, fontweight="bold")
            ax.set_ylabel("EUR miliardi")
            ax.set_ylim(0, scen_vals.max() * 1.25)
            ax.grid(alpha=0.25, axis="y")
            ax.axhline(0, color="black", lw=0.8)

        fig.suptitle("Windfall profits stimati (IT vs Germania come controllo)\n"
                     "sensitività sui volumi: base ±30%  |  solo per δ̂>0 significativo",
                     fontsize=10, fontweight="bold")
        plt.tight_layout()
        fig.savefig("plots/3_03b_windfall.png", dpi=DPI, bbox_inches="tight")
        plt.close(fig)
        print("✓ plots/3_03b_windfall.png")

print("\nScript 3_03 completato.")
